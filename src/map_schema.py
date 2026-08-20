import os
import json
import openpyxl

def normalize_unit(unit_str):
    if not unit_str:
        return ""
    u = unit_str.strip().lower()
    if u in ["cu.m", "cum", "m3", "m³", "cubic metre"]:
        return "cum"
    if u in ["sq.m", "sqm", "m2", "m²", "square metre"]:
        return "sqm"
    if u in ["r.m", "rm", "m", "metre"]:
        return "m"
    if u in ["kg", "kgs", "kilogram"]:
        return "kg"
    return unit_str

def map_data_to_excel_template():
    output_dir = os.getenv("OUTPUT_DIR", "output")
    json_path = os.path.join(output_dir, "raw_extracted.json")
    template_path = os.getenv("EXCEL_TEMPLATE_PATH", "AMP_Passport_Template.xlsx")
    final_excel_path = os.path.join(output_dir, "passport_filled.xlsx")

    if not os.path.exists(json_path):
        print(f"[!] Error: {json_path} not found.")
        return

    with open(json_path, "r", encoding="utf-8") as f:
        items = json.load(f)

    wb = openpyxl.load_workbook(template_path)
    ws = wb["Material Passport"]

    # Clear example rows starting from Row 4
    for r in range(4, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    row_idx = 4
    for item in items:
        # Col 2: BOQ Item No.
        ws.cell(row=row_idx, column=2, value=item.get("item_no"))
        # Col 5: Description
        ws.cell(row=row_idx, column=5, value=item.get("description"))
        # Col 6: Floor / Section
        ws.cell(row=row_idx, column=6, value=item.get("floor_section"))
        # Col 7: Discipline
        ws.cell(row=row_idx, column=7, value=item.get("discipline") or "Civil")
        # Col 8: Material / Product
        ws.cell(row=row_idx, column=8, value=item.get("material_product"))
        # Col 10: Material Category
        ws.cell(row=row_idx, column=10, value=item.get("material_category"))
        # Col 12: Grade
        ws.cell(row=row_idx, column=12, value=item.get("grade_mix"))
        # Col 14: Original Quantity
        ws.cell(row=row_idx, column=14, value=item.get("quantity"))
        # Col 15: Original Unit
        ws.cell(row=row_idx, column=15, value=normalize_unit(item.get("unit")))
        # Col 27: Schedule (DSR/SOR)
        ws.cell(row=row_idx, column=27, value="DSR" if item.get("dsr_code") else None)
        # Col 28: Schedule Item Code (FIXED MAPPING)
        ws.cell(row=row_idx, column=28, value=item.get("dsr_code"))
        # Col 47: Unit Rate
        ws.cell(row=row_idx, column=47, value=item.get("rate"))
        # Col 48: Total Cost
        ws.cell(row=row_idx, column=48, value=item.get("amount"))
        # Col 49: Currency
        ws.cell(row=row_idx, column=49, value="INR")

        row_idx += 1

    try:
        wb.save(final_excel_path)
        print(f"[✓] Successfully populated {len(items)} items into '{final_excel_path}'")
    except PermissionError:
        fallback = os.path.join(output_dir, "passport_filled_updated.xlsx")
        wb.save(fallback)
        print(f"[!] Primary file was open. Saved to fallback: '{fallback}'")

if __name__ == "__main__":
    map_data_to_excel_template()